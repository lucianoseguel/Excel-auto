#organizador de planillas de Western union
import pandas as pd
import datetime
import openpyxl 
from openpyxl  import load_workbook
from datetime import datetime, timedelta
import os
import time
import csv


def main():
    swi = 1
    while swi == 1:
       
            os.system('cls')
            print('----     MENU PARA HACER LA CAJA DE WESTERN UNION ')
            selector = input('--------- CREAR PLANILLA (1) CERRAR LA CAJA (2): ')

            if selector == '1':
                creaplanilla()
            elif selector == '2':
                hacerlacaja()
                swi = 0




#Hacer y cerrar la caja
def hacerlacaja():

    try:
        print('Buscando planilla')
        time.sleep(20)
        
        

        hoy = extrae_fecha()
        fecha_file_csv_today = extrae_fecha_paraCSV()
        year = extrae_year()
        filename_today = f'G:\Mi unidad\{str(year)} Alem\{str(hoy)}.xlsx'
        # Cargar el archivo Excel
        planilla_nueva = openpyxl.load_workbook(filename_today,data_only=False)
        # Seleccionar una hoja de trabajo
                
        planilla_nueva.active
        wester = planilla_nueva['Western']

        


        
        #Abrir CSV
        rutacsv = f'C:/Users/Admin/Downloads/Transacciones_{fecha_file_csv_today}_A_{fecha_file_csv_today}.csv'
        #----------------------------Envios----------------------------------
        #rutacsv = f'C:/Users/lucia/Downloads/Transacciones_{fecha_file_csv_today}_A_{fecha_file_csv_today}.csv'
        envios = []
        
        with open((rutacsv),mode='r') as file:
            csv_file = csv.reader(file)
            


            for i in range(4):
                next(csv_file) 
                
            os.system('cls')
            print('Imprimiendo datos')
            time.sleep(1)

            for i in csv_file:
                try:
                    print((i[19]))
                    envios.append(float(i[19]))
                
                except IndexError:
                    break

            start_row = 6  # Ajusta según sea necesario
            column = 'A'  # Ajusta según sea necesario

            for row_index, value in enumerate(envios, start=start_row):
                wester.cell(row=row_index, column=5, value=value)

        #----------------------------PAGOS----------------------------------

        pagos1 = []
        pagos2 = []
        mtcn1 = []
        mtcn2 = []

        with open((rutacsv),mode='r') as file:
            csv_file = csv.reader(file)
            


            for i in range(4 + len(envios) + 19):
                next(csv_file) 
                

            cantidad_pagos = 0
            for i in csv_file:
                cantidad_pagos += 1
                try:
                    if cantidad_pagos > 45: #Creo listas para almacenar los datos de los pagos
                        
                        print((i[21]))
                        pagos1.append(float(i[21]))
                        mtcn1.append(int(i[6]))
                    else:
                        print((i[21]))
                        pagos2.append(float(i[21]))
                        mtcn2.append(int(i[6]))


            
                
                except IndexError: #Cuando la celda esta en blanco hago el break para que no me salte el error de la celda en blanco
                    break
        
            os.system('cls')
            print('Copiando datos')
            time.sleep(1)
            

            start_row = 6  # Ajusta según sea necesario
            column = 'A'  # Ajusta según sea necesario

            for row_index, value in enumerate(pagos2, start=start_row): #Pega los pagos en la columna de la izquierda
                wester.cell(row=row_index, column=2, value=value)

            for row_index, value in enumerate(pagos1, start=start_row): #Pega los pagos en la columna de la derecha
                wester.cell(row=row_index, column=11, value=value)


            #-----------------------MTCN-----------------------


            for row_index, value in enumerate(mtcn2, start=start_row): #Pega los mtcn en la columna de la izquierda
                wester.cell(row=row_index, column=3, value=value)

            for row_index, value in enumerate(mtcn1, start=start_row): #Pega los mtcn en la columna de la derecha
                wester.cell(row=row_index, column=12, value=value)

            os.system('cls')
            print('Pegando datos')
            time.sleep(1)
            





        #Guardado final
        planilla_nueva.save(filename_today)
        planilla_nueva.close()
        os.system('cls')
        print('Indexado con exito CERRANDO PROGRAMA')
        time.sleep(4)

    except FileNotFoundError:
        print("Archivo no encontrado")
        time.sleep(5)

    except IndexError:
        print("Error en lectura de archivo")
        input("")
    
    except Exception: 
        print("Error desconocido")
        input("")









# SHEET CREATOR
def creaplanilla():
    # Ruta al archivo Excel
    swi =True
    while swi == True:
        try:
            os.system('cls')
            print('----     CREADOR DE PLANILLAS DIARIAS (usar solo al principio del dia 1 vez, sino cerrar)\n\n')
            
            hoy = extrae_fecha()
            ayer = extrae_ayer()
            hoy2 = extrae_fecha_plan()
            year = extrae_year()
            print(f'De {ayer} ---->  {hoy}')
            
         #filename_yesterday = (f'G:\Mi unidad\ 2024\{str(ayer)}.xlsx')
            filename_yesterday = (f'G:\Mi unidad\{str(year)} Alem\{str(ayer)}.xlsx')

            
                

            # Cargar el archivo Excel
            planilla_nueva = openpyxl.load_workbook(filename_yesterday,data_only=True)
            # Seleccionar una hoja de trabajo
            
            planilla_nueva.active
            wester = planilla_nueva['Western']

            

        
        
            # Modificar celdas
            saldo_inicial = saldo(wester)
            

            planilla_nueva.close()


            # Cargar el archivo Excel
            planilla_nueva = openpyxl.load_workbook(filename_yesterday,data_only=False)
            # Seleccionar una hoja de trabajo
            
            planilla_nueva.active
            wester = planilla_nueva['Western']

            borrador_tabla(wester,'B6:E50') #Primera tabla
            borrador_tabla(wester,'K6:L50') #Segunda tabla tabla
            borrador_tabla(wester,'B55:D68') #Tabla abajo
            borrador_tabla(wester,'E57:E68') #T1 y t2'''

            wester['C4'] = float(saldo_inicial)
            wester['F3'] = (hoy2)


            # Guardar los cambios
            
            planilla_nueva.save(f'G:\Mi unidad\{str(year)} Alem\{str(hoy)}.xlsx')
            planilla_nueva.close()
            swi = False
        except FileNotFoundError:
            print(f'File {ayer} no exist ')
            time.sleep(3)



def borrador_tabla(hoja,rango): #Borra los datos de la tabla
    
    rango1 = hoja[rango]

    for fila in rango1:
        for celda in fila:
            celda.value = ''



def extrae_fecha(): #Extrae la fecha de hoy
    fecha = datetime.now()
    formateada = fecha.strftime('%d-%m-%Y')
    return formateada

def extrae_fecha_plan(): #Ecxtrae la fecha de hoy para la planilla
    fecha = datetime.now()
    formateada = fecha.strftime('%d/%m/%Y')
    return formateada

def extrae_ayer(): #Extrae la fecha del dia anterior a indexar
    select = input('Dias a indexar atras: ')
    while not select.isdigit():
        select = input('Dias a indexar atras: ')

    ayer =  datetime.now()  - timedelta(days=int(select))
    fomrateada = ayer.strftime('%d-%m-%Y')
    return(fomrateada)

def extrae_year(): #Devuelve el año actual
    
    hoy =  datetime.now() 
    year = hoy.strftime('%Y')
    return(year)



def saldo(hoja):#Copia y pega el saldo a la casilla de saldo inicial
 
    saldo =  hoja['I3'].value
    
    return saldo


def extrae_fecha_paraCSV(): #Extrae la fecha para indexar en el CSV
    fecha = datetime.now()
    para_csv = fecha.strftime('%Y%m%d')
    return para_csv
   

if __name__ == "__main__":
    main()
